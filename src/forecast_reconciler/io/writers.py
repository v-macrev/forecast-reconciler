from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable
from zipfile import ZIP_DEFLATED, ZipFile

import polars as pl
from openpyxl import Workbook

from forecast_reconciler.exceptions import DataValidationError


ExportProgressCallback = Callable[[str], None]


@dataclass(frozen=True, slots=True)
class ExcelExportResult:
    """
    Result of workbook export.

    Attributes
    ----------
    output_path:
        Absolute path to the written workbook.
    sheet_names:
        Names of worksheets written to the workbook.
    """

    output_path: Path
    sheet_names: tuple[str, ...]


@dataclass(frozen=True, slots=True)
class ZipCsvExportResult:
    """
    Result of ZIP-based CSV export.

    Attributes
    ----------
    output_path:
        Absolute path to the written ZIP archive.
    file_names:
        Names of CSV files written into the archive.
    """

    output_path: Path
    file_names: tuple[str, ...]


def export_reconciliation_workbook(
    output_path: str | Path,
    final_allocations_df: pl.DataFrame,
    group_summary_df: pl.DataFrame,
    sku_variance_df: pl.DataFrame,
    integrity_summary_df: pl.DataFrame,
    progress_callback: ExportProgressCallback | None = None,
) -> ExcelExportResult:
    """
    Export reconciliation artefacts into a deterministic Excel workbook.

    Uses OpenPyXL write-only mode for faster large-sheet export.
    """
    _validate_export_inputs(
        final_allocations_df=final_allocations_df,
        group_summary_df=group_summary_df,
        sku_variance_df=sku_variance_df,
        integrity_summary_df=integrity_summary_df,
    )

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    _emit_progress(progress_callback, "Creating workbook")

    workbook = Workbook(write_only=True)

    sheet_payloads = [
        ("final_allocations", final_allocations_df),
        ("group_summary", group_summary_df),
        ("sku_variance", sku_variance_df),
        ("integrity_summary", integrity_summary_df),
    ]

    written_sheet_names: list[str] = []

    for sheet_name, df in sheet_payloads:
        _emit_progress(progress_callback, f"Writing sheet: {sheet_name}")
        worksheet = workbook.create_sheet(title=sheet_name)
        _write_dataframe_to_worksheet(worksheet=worksheet, df=df)
        written_sheet_names.append(sheet_name)

    _emit_progress(progress_callback, "Saving workbook")
    workbook.save(path)

    _emit_progress(progress_callback, "Workbook export completed")

    return ExcelExportResult(
        output_path=path,
        sheet_names=tuple(written_sheet_names),
    )


def export_reconciliation_csv_zip(
    output_path: str | Path,
    final_allocations_df: pl.DataFrame,
    group_summary_df: pl.DataFrame,
    sku_variance_df: pl.DataFrame,
    integrity_summary_df: pl.DataFrame,
    progress_callback: ExportProgressCallback | None = None,
) -> ZipCsvExportResult:
    """
    Export reconciliation artefacts into a ZIP archive containing CSV files.
    """
    _validate_export_inputs(
        final_allocations_df=final_allocations_df,
        group_summary_df=group_summary_df,
        sku_variance_df=sku_variance_df,
        integrity_summary_df=integrity_summary_df,
    )

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)

    file_payloads = [
        ("final_allocations.csv", final_allocations_df),
        ("group_summary.csv", group_summary_df),
        ("sku_variance.csv", sku_variance_df),
        ("integrity_summary.csv", integrity_summary_df),
    ]

    _emit_progress(progress_callback, "Creating ZIP archive")

    written_file_names: list[str] = []

    with ZipFile(path, mode="w", compression=ZIP_DEFLATED) as zip_file:
        for file_name, df in file_payloads:
            _emit_progress(progress_callback, f"Writing file: {file_name}")
            csv_bytes = _dataframe_to_csv_bytes(df)
            zip_file.writestr(file_name, csv_bytes)
            written_file_names.append(file_name)

    _emit_progress(progress_callback, "CSV ZIP export completed")

    return ZipCsvExportResult(
        output_path=path,
        file_names=tuple(written_file_names),
    )


def _validate_export_inputs(
    final_allocations_df: pl.DataFrame,
    group_summary_df: pl.DataFrame,
    sku_variance_df: pl.DataFrame,
    integrity_summary_df: pl.DataFrame,
) -> None:
    datasets = {
        "final_allocations": final_allocations_df,
        "group_summary": group_summary_df,
        "sku_variance": sku_variance_df,
        "integrity_summary": integrity_summary_df,
    }

    for dataset_name, df in datasets.items():
        if not isinstance(df, pl.DataFrame):
            raise DataValidationError(
                f"Export input '{dataset_name}' must be a Polars DataFrame."
            )

        if df.width == 0:
            raise DataValidationError(
                f"Export input '{dataset_name}' must contain at least one column."
            )


def _write_dataframe_to_worksheet(worksheet: Any, df: pl.DataFrame) -> None:
    """
    Write a Polars DataFrame into an OpenPyXL worksheet, including header row.
    """
    worksheet.append(list(df.columns))

    for row in df.iter_rows(named=False):
        worksheet.append([_serialise_cell_value(value) for value in row])


def _serialise_cell_value(value: Any) -> Any:
    """
    Convert Python/Polars values into Excel-friendly cell values.
    """
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    return value


def _dataframe_to_csv_bytes(df: pl.DataFrame) -> bytes:
    """
    Serialise a Polars DataFrame to UTF-8 CSV bytes.
    """
    buffer = StringIO()
    df.write_csv(buffer)
    return buffer.getvalue().encode("utf-8")


def _emit_progress(
    progress_callback: ExportProgressCallback | None,
    message: str,
) -> None:
    if progress_callback is not None:
        progress_callback(message)