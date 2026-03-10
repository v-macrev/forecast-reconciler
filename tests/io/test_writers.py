from datetime import date
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from forecast_reconciler.exceptions import DataValidationError
from forecast_reconciler.io.writers import (
    ExcelExportResult,
    export_reconciliation_workbook,
)


def test_export_reconciliation_workbook_writes_expected_sheets(tmp_path: Path):
    output_file = tmp_path / "reconciliation_output.xlsx"

    final_allocations_df = pl.DataFrame(
        {
            "period": [date(2026, 1, 1)],
            "market": ["SP"],
            "channel": ["Retail"],
            "sku": ["SKU-001"],
            "final_allocated_qty": [100.0],
        }
    )

    group_summary_df = pl.DataFrame(
        {
            "period": [date(2026, 1, 1)],
            "market": ["SP"],
            "channel": ["Retail"],
            "macro_target_qty": [100.0],
            "final_allocated_group_qty": [100.0],
        }
    )

    sku_variance_df = pl.DataFrame(
        {
            "period": [date(2026, 1, 1)],
            "market": ["SP"],
            "channel": ["Retail"],
            "sku": ["SKU-001"],
            "baseline_qty": [80.0],
            "final_allocated_qty": [100.0],
        }
    )

    integrity_summary_df = pl.DataFrame(
        {
            "validated_group_count": [1],
            "groups_with_gap_count": [0],
            "negative_allocation_count": [0],
            "unmatched_macro_group_count": [0],
            "unmatched_granular_group_count": [0],
            "is_valid": [True],
        }
    )

    result = export_reconciliation_workbook(
        output_path=output_file,
        final_allocations_df=final_allocations_df,
        group_summary_df=group_summary_df,
        sku_variance_df=sku_variance_df,
        integrity_summary_df=integrity_summary_df,
    )

    assert isinstance(result, ExcelExportResult)
    assert result.output_path == output_file.resolve()
    assert result.sheet_names == (
        "final_allocations",
        "group_summary",
        "sku_variance",
        "integrity_summary",
    )
    assert output_file.exists()

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == [
        "final_allocations",
        "group_summary",
        "sku_variance",
        "integrity_summary",
    ]


def test_export_reconciliation_workbook_writes_headers_and_rows(tmp_path: Path):
    output_file = tmp_path / "reconciliation_output.xlsx"

    final_allocations_df = pl.DataFrame(
        {
            "period": [date(2026, 1, 1)],
            "market": ["SP"],
            "channel": ["Retail"],
            "sku": ["SKU-001"],
            "final_allocated_qty": [100.0],
        }
    )

    group_summary_df = pl.DataFrame({"metric": ["ok"], "value": [1]})
    sku_variance_df = pl.DataFrame({"sku": ["SKU-001"], "delta": [20.0]})
    integrity_summary_df = pl.DataFrame({"is_valid": [True]})

    export_reconciliation_workbook(
        output_path=output_file,
        final_allocations_df=final_allocations_df,
        group_summary_df=group_summary_df,
        sku_variance_df=sku_variance_df,
        integrity_summary_df=integrity_summary_df,
    )

    workbook = load_workbook(output_file)
    worksheet = workbook["final_allocations"]

    header_row = [cell.value for cell in worksheet[1]]
    data_row = [cell.value for cell in worksheet[2]]

    assert header_row == [
        "period",
        "market",
        "channel",
        "sku",
        "final_allocated_qty",
    ]
    assert data_row[0].date() == date(2026, 1, 1)
    assert data_row[1:] == ["SP", "Retail", "SKU-001", 100.0]


def test_export_reconciliation_workbook_rejects_non_dataframe_inputs(tmp_path: Path):
    output_file = tmp_path / "reconciliation_output.xlsx"

    valid_df = pl.DataFrame({"col": [1]})

    try:
        export_reconciliation_workbook(
            output_path=output_file,
            final_allocations_df=valid_df,
            group_summary_df=valid_df,
            sku_variance_df=valid_df,
            integrity_summary_df={"is_valid": [True]},
        )
    except DataValidationError as exc:
        assert str(exc) == (
            "Export input 'integrity_summary' must be a Polars DataFrame."
        )
    else:
        raise AssertionError("Expected DataValidationError for non-DataFrame export input.")


def test_export_reconciliation_workbook_rejects_zero_column_inputs(tmp_path: Path):
    output_file = tmp_path / "reconciliation_output.xlsx"

    empty_column_df = pl.DataFrame()

    try:
        export_reconciliation_workbook(
            output_path=output_file,
            final_allocations_df=empty_column_df,
            group_summary_df=pl.DataFrame({"col": [1]}),
            sku_variance_df=pl.DataFrame({"col": [1]}),
            integrity_summary_df=pl.DataFrame({"col": [1]}),
        )
    except DataValidationError as exc:
        assert str(exc) == (
            "Export input 'final_allocations' must contain at least one column."
        )
    else:
        raise AssertionError("Expected DataValidationError for zero-column export input.")