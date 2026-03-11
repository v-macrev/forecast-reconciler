from datetime import date
from pathlib import Path

import polars as pl
from openpyxl import load_workbook

from forecast_reconciler.config import ReconciliationConfig
from forecast_reconciler.io.writers import export_reconciliation_workbook
from forecast_reconciler.normalisation.standardise import (
    standardise_granular_input,
    standardise_macro_input,
)
from forecast_reconciler.reconciliation.allocator import redistribute_macro_targets
from forecast_reconciler.reconciliation.rounding import apply_deterministic_rounding
from forecast_reconciler.reconciliation.weights import calculate_weights
from forecast_reconciler.reporting.summaries import build_reporting_views
from forecast_reconciler.validation.integrity import validate_reconciliation_integrity


def test_end_to_end_reconciliation_pipeline(tmp_path: Path):
    config = ReconciliationConfig(
        group_keys=("period", "market", "channel"),
        quantity_mode="integer",
        quantity_decimals=0,
        zero_baseline_mode="fail",
        allow_negative_allocations=False,
        enforce_exact_totals=True,
    )

    macro_raw = pl.DataFrame(
        {
            "period": ["2026-01", "2026-01", "2026-02", "2026-02"],
            "market": ["SP", "RJ", "SP", "RJ"],
            "channel": ["Retail", "Retail", "Retail", "Retail"],
            "macro_target_qty": [120, 50, 90, 80],
        }
    )

    granular_raw = pl.DataFrame(
        {
            "period": [
                "2026-01",
                "2026-01",
                "2026-01",
                "2026-01",
                "2026-02",
                "2026-02",
                "2026-02",
                "2026-02",
            ],
            "market": ["SP", "SP", "RJ", "RJ", "SP", "SP", "RJ", "RJ"],
            "channel": ["Retail"] * 8,
            "sku": [
                "SKU-001",
                "SKU-002",
                "SKU-010",
                "SKU-011",
                "SKU-001",
                "SKU-002",
                "SKU-010",
                "SKU-011",
            ],
            "baseline_qty": [60, 40, 20, 30, 30, 70, 50, 30],
        }
    )

    macro_std = standardise_macro_input(df=macro_raw, config=config)
    granular_std = standardise_granular_input(df=granular_raw, config=config)

    assert macro_std.columns == ["period", "market", "channel", "macro_target_qty"]
    assert granular_std.columns == ["period", "market", "channel", "sku", "baseline_qty"]

    weight_result = calculate_weights(granular_df=granular_std, config=config)
    assert weight_result.zero_baseline_groups.height == 0

    redistribution_result = redistribute_macro_targets(
        macro_df=macro_std,
        weighted_granular_df=weight_result.weights,
        config=config,
    )

    assert redistribution_result.unmatched_macro_groups.height == 0
    assert redistribution_result.unmatched_granular_groups.height == 0

    rounding_result = apply_deterministic_rounding(
        allocation_df=redistribution_result.allocations,
        config=config,
    )

    rounded_allocations = rounding_result.allocations.sort(["period", "market", "sku"])

    expected_final_allocations = [
        {
            "period": date(2026, 1, 1),
            "market": "RJ",
            "channel": "Retail",
            "sku": "SKU-010",
            "final_allocated_qty": 20.0,
        },
        {
            "period": date(2026, 1, 1),
            "market": "RJ",
            "channel": "Retail",
            "sku": "SKU-011",
            "final_allocated_qty": 30.0,
        },
        {
            "period": date(2026, 1, 1),
            "market": "SP",
            "channel": "Retail",
            "sku": "SKU-001",
            "final_allocated_qty": 72.0,
        },
        {
            "period": date(2026, 1, 1),
            "market": "SP",
            "channel": "Retail",
            "sku": "SKU-002",
            "final_allocated_qty": 48.0,
        },
        {
            "period": date(2026, 2, 1),
            "market": "RJ",
            "channel": "Retail",
            "sku": "SKU-010",
            "final_allocated_qty": 50.0,
        },
        {
            "period": date(2026, 2, 1),
            "market": "RJ",
            "channel": "Retail",
            "sku": "SKU-011",
            "final_allocated_qty": 30.0,
        },
        {
            "period": date(2026, 2, 1),
            "market": "SP",
            "channel": "Retail",
            "sku": "SKU-001",
            "final_allocated_qty": 27.0,
        },
        {
            "period": date(2026, 2, 1),
            "market": "SP",
            "channel": "Retail",
            "sku": "SKU-002",
            "final_allocated_qty": 63.0,
        },
    ]

    assert rounded_allocations.select(
        ["period", "market", "channel", "sku", "final_allocated_qty"]
    ).to_dicts() == expected_final_allocations

    integrity_result = validate_reconciliation_integrity(
        rounded_allocations_df=rounding_result.allocations,
        unmatched_macro_groups=redistribution_result.unmatched_macro_groups,
        unmatched_granular_groups=redistribution_result.unmatched_granular_groups,
        config=config,
    )

    assert integrity_result.is_valid is True
    assert integrity_result.summary.to_dicts() == [
        {
            "validated_group_count": 4,
            "groups_with_gap_count": 0,
            "negative_allocation_count": 0,
            "unmatched_macro_group_count": 0,
            "unmatched_granular_group_count": 0,
            "is_valid": True,
        }
    ]

    reporting_result = build_reporting_views(
        rounded_allocations_df=rounding_result.allocations,
        config=config,
    )

    group_summary = reporting_result.group_summary.sort(["period", "market"])

    assert group_summary.select(
        [
            "period",
            "market",
            "channel",
            "macro_target_qty",
            "baseline_group_qty",
            "final_allocated_group_qty",
            "final_gap_to_target",
        ]
    ).to_dicts() == [
        {
            "period": date(2026, 1, 1),
            "market": "RJ",
            "channel": "Retail",
            "macro_target_qty": 50.0,
            "baseline_group_qty": 50.0,
            "final_allocated_group_qty": 50.0,
            "final_gap_to_target": 0.0,
        },
        {
            "period": date(2026, 1, 1),
            "market": "SP",
            "channel": "Retail",
            "macro_target_qty": 120.0,
            "baseline_group_qty": 100.0,
            "final_allocated_group_qty": 120.0,
            "final_gap_to_target": 0.0,
        },
        {
            "period": date(2026, 2, 1),
            "market": "RJ",
            "channel": "Retail",
            "macro_target_qty": 80.0,
            "baseline_group_qty": 80.0,
            "final_allocated_group_qty": 80.0,
            "final_gap_to_target": 0.0,
        },
        {
            "period": date(2026, 2, 1),
            "market": "SP",
            "channel": "Retail",
            "macro_target_qty": 90.0,
            "baseline_group_qty": 100.0,
            "final_allocated_group_qty": 90.0,
            "final_gap_to_target": 0.0,
        },
    ]

    output_file = tmp_path / "forecast_reconciliation.xlsx"

    export_result = export_reconciliation_workbook(
        output_path=output_file,
        final_allocations_df=rounding_result.allocations,
        group_summary_df=reporting_result.group_summary,
        sku_variance_df=reporting_result.sku_variance,
        integrity_summary_df=integrity_result.summary,
    )

    assert export_result.output_path == output_file.resolve()
    assert export_result.sheet_names == (
        "final_allocations",
        "group_summary",
        "sku_variance",
        "integrity_summary",
    )
    assert output_file.exists()

    workbook = load_workbook(output_file)
    assert workbook.sheetnames == [
        "final_allocations",
        "group_summary",
        "sku_variance",
        "integrity_summary",
    ]

    integrity_sheet = workbook["integrity_summary"]
    integrity_header = [cell.value for cell in integrity_sheet[1]]
    integrity_values = [cell.value for cell in integrity_sheet[2]]

    assert integrity_header == [
        "validated_group_count",
        "groups_with_gap_count",
        "negative_allocation_count",
        "unmatched_macro_group_count",
        "unmatched_granular_group_count",
        "is_valid",
    ]
    assert integrity_values == [4, 0, 0, 0, 0, True]